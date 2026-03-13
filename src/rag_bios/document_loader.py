import os
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile

from langchain_core.documents import Document
from openpyxl import load_workbook
from unstructured.partition.auto import partition


LOGGER = logging.getLogger(__name__)
SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".xlsx"}
TEXT_ENCODINGS = ("utf-8", "utf-8-sig", "cp1252", "latin-1")
TXT_HEADING_PATTERN = re.compile(r"^\d+[.)]\s+")


@dataclass(slots=True)
class DocumentLoadResult:
    documents: list[Document]
    processed_files: list[str]
    warnings: list[str]


@dataclass(slots=True)
class TxtSection:
    content: str
    title: str | None
    command_lines: list[str]


def save_uploaded_file(uploaded_file) -> Path:
    suffix = Path(uploaded_file.name).suffix.lower()
    with NamedTemporaryFile(delete=False, suffix=suffix) as temp_file:
        temp_file.write(uploaded_file.getbuffer())
        return Path(temp_file.name)


def load_uploaded_documents(uploaded_files, document_languages: list[str] | None = None) -> DocumentLoadResult:
    documents: list[Document] = []
    processed_files: list[str] = []
    warnings: list[str] = []
    temp_paths: list[Path] = []

    try:
        for uploaded_file in uploaded_files:
            temp_path = save_uploaded_file(uploaded_file)
            temp_paths.append(temp_path)

            try:
                loaded_documents = load_file(
                    temp_path,
                    uploaded_file.name,
                    document_languages=document_languages,
                )
            except Exception as exc:
                LOGGER.exception("No se pudo procesar el archivo %s", uploaded_file.name)
                warnings.append(f"No pude procesar {uploaded_file.name}: {exc}")
                continue

            if not loaded_documents:
                warnings.append(
                    f"Omiti {uploaded_file.name} porque no extraje contenido util para indexar."
                )
                continue

            documents.extend(loaded_documents)
            processed_files.append(uploaded_file.name)
    finally:
        for temp_path in temp_paths:
            if temp_path.exists():
                os.remove(temp_path)

    return DocumentLoadResult(
        documents=documents,
        processed_files=processed_files,
        warnings=warnings,
    )


def load_file(
    file_path: Path,
    display_name: str | None = None,
    *,
    document_languages: list[str] | None = None,
) -> list[Document]:
    suffix = file_path.suffix.lower()
    source_name = display_name or file_path.name

    if suffix not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Formato no soportado: {suffix}")

    if suffix == ".txt":
        return _load_txt(file_path, source_name)

    if suffix == ".xlsx":
        return _load_xlsx(file_path, source_name)

    elements = partition(
        filename=str(file_path),
        languages=document_languages,
    )
    documents: list[Document] = []

    for index, element in enumerate(elements, start=1):
        text = getattr(element, "text", "")
        if not text or not text.strip():
            continue

        metadata = {
            "source": source_name,
            "element_index": index,
            "file_type": suffix.lstrip("."),
        }
        page_number = getattr(getattr(element, "metadata", None), "page_number", None)
        if page_number is not None:
            metadata["page_number"] = page_number

        documents.append(Document(page_content=text.strip(), metadata=metadata))

    return documents


def _load_txt(file_path: Path, source_name: str) -> list[Document]:
    raw_text = _read_text_file(file_path)
    sections = _split_txt_sections(raw_text)
    documents: list[Document] = []

    for index, section in enumerate(sections, start=1):
        metadata = {
            "source": source_name,
            "element_index": index,
            "file_type": "txt",
        }
        if section.title:
            metadata["section_title"] = section.title
        if section.command_lines:
            metadata["command_lines"] = section.command_lines

        documents.append(
            Document(
                page_content=section.content,
                metadata=metadata,
            )
        )

    return documents


def _load_xlsx(file_path: Path, source_name: str) -> list[Document]:
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    documents: list[Document] = []

    try:
        for sheet in workbook.worksheets:
            populated_rows = [
                (row_index, [_normalize_cell(value) for value in row])
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1)
            ]
            populated_rows = [
                (row_index, values)
                for row_index, values in populated_rows
                if any(values)
            ]

            if not populated_rows:
                continue

            uses_header_row = len(populated_rows) >= 2
            header_row = populated_rows[0][1] if uses_header_row else []
            data_rows = populated_rows[1:] if uses_header_row else populated_rows
            header_names = _build_header_names(header_row, data_rows)

            for row_index, values in data_rows:
                # Keep each row intact so exact date/column questions can be answered
                # without losing the relationship between fields during chunking.
                row_pairs = []
                date_aliases = []
                row_data: dict[str, str] = {}
                for column_index, value in enumerate(values, start=1):
                    if not value:
                        continue
                    header_name = _header_name_for_column(header_names, column_index)
                    row_pairs.append(f"{header_name}: {value}")
                    row_data[header_name] = value
                    if header_name.strip().lower() == "date":
                        date_aliases.append(f"Fecha: {value}")

                if not row_pairs:
                    continue

                content = "\n".join(
                    [f"Hoja: {sheet.title}", f"Fila: {row_index}", *date_aliases, *row_pairs]
                )
                documents.append(
                    Document(
                        page_content=content,
                        metadata={
                            "source": source_name,
                            "sheet_name": sheet.title,
                            "row_number": row_index,
                            "file_type": "xlsx",
                            "column_names": header_names,
                            "row_data": row_data,
                            "date_value": row_data.get("Date") or row_data.get("date"),
                        },
                    )
                )
    finally:
        workbook.close()

    return documents


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text


def _read_text_file(file_path: Path) -> str:
    file_bytes = file_path.read_bytes()
    for encoding in TEXT_ENCODINGS:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _split_txt_sections(raw_text: str) -> list[TxtSection]:
    lines = [
        _normalize_text_line(line)
        for line in raw_text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    ]

    sections: list[TxtSection] = []
    pending_context: list[str] = []
    current_prefix: list[str] = []
    current_heading: str | None = None
    current_body: list[str] = []

    for line in lines:
        if not line:
            continue

        if _is_txt_heading(line):
            if current_heading is not None:
                if current_body:
                    sections.append(_compose_txt_section(current_prefix, current_heading, current_body))
                else:
                    pending_context.extend(current_prefix)
                    pending_context.append(current_heading)
                current_prefix = []
                current_body = []

            current_prefix = pending_context
            pending_context = []
            current_heading = line
            continue

        if current_heading is not None:
            current_body.append(line)
        else:
            pending_context.append(line)

    if current_heading is not None:
        if current_body:
            sections.append(_compose_txt_section(current_prefix, current_heading, current_body))
        else:
            pending_context.extend(current_prefix)
            pending_context.append(current_heading)

    if pending_context:
        sections.insert(
            0,
            TxtSection(
                content="\n".join(pending_context),
                title=None,
                command_lines=_extract_command_lines(pending_context),
            ),
        )

    return [section for section in sections if section.content.strip()]


def _normalize_text_line(line: str) -> str:
    normalized = line.strip()
    normalized = normalized.replace("•", "-")
    normalized = re.sub(r"\s+", " ", normalized)
    return normalized


def _is_txt_heading(line: str) -> bool:
    if line.startswith("- "):
        return False
    if TXT_HEADING_PATTERN.match(line):
        return True
    return line.endswith(":") and len(line) <= 120


def _compose_txt_section(prefix: list[str], heading: str, body: list[str]) -> TxtSection:
    section_lines = [*prefix, heading, *body]
    return TxtSection(
        content="\n".join(section_lines),
        title=heading,
        command_lines=_extract_command_lines(body),
    )


def _extract_command_lines(lines: list[str]) -> list[str]:
    return [
        line
        for line in lines
        if re.match(r"^(docker|ollama|\/set)\b", line, re.IGNORECASE)
    ]


def _build_header_names(
    header_row: list[str],
    data_rows: list[tuple[int, list[str]]],
) -> list[str]:
    max_columns = max(
        [len(header_row), *[len(values) for _, values in data_rows]],
        default=0,
    )
    unique_headers: list[str] = []
    seen_headers: dict[str, int] = {}

    for column_index in range(1, max_columns + 1):
        raw_header = header_row[column_index - 1].strip() if column_index <= len(header_row) else ""
        header_name = raw_header or f"col_{column_index}"
        duplicate_index = seen_headers.get(header_name, 0)
        if duplicate_index:
            header_name = f"{header_name}_{duplicate_index + 1}"
        seen_headers[raw_header or f"col_{column_index}"] = duplicate_index + 1
        unique_headers.append(header_name)

    return unique_headers


def _header_name_for_column(header_names: list[str], column_index: int) -> str:
    if column_index <= len(header_names):
        return header_names[column_index - 1]
    return f"col_{column_index}"
